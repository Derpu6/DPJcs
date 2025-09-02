import streamlit as st
from utils import qa_agent, QwenChat, QwenEmbeddings
from langchain.memory import ConversationBufferMemory
from langchain_core.messages import HumanMessage, AIMessage
import logging
import traceback
import os
from PIL import Image
import time
from typing import Dict, Any, Optional, List, Tuple
import pandas as pd
from datetime import datetime
import shutil
import threading
import uuid
import json

# ==================== 初始化设置 ====================
st.set_page_config(
    page_title="智能单片机问答工具",
    page_icon="🖥️",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# 初始化日志
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


# ==================== 高分问题管理器 ====================
class TopQuestionManager:
    """管理高分问题的JSON存储"""

    def __init__(self, storage_dir="top_questions"):
        self.storage_dir = os.path.abspath(storage_dir)
        os.makedirs(self.storage_dir, exist_ok=True)
        self.questions = []  # 格式: [{"id": str, "question": str, "answer": str}]
        self._load_questions()

    def _load_questions(self):
        """从JSON文件加载所有高分问题"""
        self.questions = []
        if not os.path.exists(self.storage_dir):
            logger.warning(f"高分问题目录不存在: {self.storage_dir}")
            return

        for filename in os.listdir(self.storage_dir):
            if filename.endswith(".json"):
                filepath = os.path.join(self.storage_dir, filename)
                try:
                    with open(filepath, "r", encoding="utf-8") as f:
                        data = json.load(f)
                        self.questions.append(data)
                    logger.info(f"成功加载高分问题: {filename}")
                except Exception as e:
                    logger.error(f"加载高分问题文件失败 {filename}: {str(e)}")
                    logger.error(traceback.format_exc())

    def save_question(self, question_id: str, question: str, answer: str):
        """保存高分问题到JSON文件"""
        try:
            # 创建问题数据
            question_data = {
                "id": question_id,
                "question": question,
                "answer": answer,
                "timestamp": datetime.now().isoformat()
            }

            # 生成文件名
            filename = f"question_{question_id}.json"
            filepath = os.path.join(self.storage_dir, filename)

            # 保存到文件
            with open(filepath, "w", encoding="utf-8") as f:
                json.dump(question_data, f, ensure_ascii=False, indent=2)

            # 更新内存中的列表
            self.questions.append(question_data)
            logger.info(f"高分问题保存成功: {filename}")
            return True
        except Exception as e:
            logger.error(f"保存高分问题失败: {str(e)}")
            logger.error(traceback.format_exc())
            return False

    def get_all_questions(self) -> List[Dict[str, Any]]:
        """获取所有高分问题"""
        return self.questions

    def get_question_by_id(self, question_id: str) -> Optional[Dict[str, Any]]:
        """通过ID获取高分问题"""
        for q in self.questions:
            if q["id"] == question_id:
                return q
        return None


# ==================== 统计模块 ====================
class ExcelTracker:
    def __init__(self):
        self.excel_path = os.path.abspath("学习记录.xlsx")
        self._lock = threading.Lock()
        self._backup_dir = os.path.abspath("backup")
        self.top_question_manager = TopQuestionManager()  # 添加高分问题管理器

        # 确保目录存在
        os.makedirs(self._backup_dir, exist_ok=True)
        os.makedirs(os.path.dirname(self.excel_path), exist_ok=True)

        self._init_excel_file()

    def _init_excel_file(self):
        """确保Excel文件存在并初始化"""
        with self._lock:
            if not os.path.exists(self.excel_path):
                logger.info("Excel文件不存在，正在初始化...")
                try:
                    self._create_new_excel_file()
                    logger.info("Excel文件初始化成功")
                except Exception as e:
                    logger.error(f"Excel文件初始化失败: {str(e)}")
                    raise
            else:
                logger.info("Excel文件已存在")

    def _create_new_excel_file(self):
        """创建全新的Excel文件"""
        try:
            with pd.ExcelWriter(self.excel_path, engine='openpyxl') as writer:
                # 学习记录表
                pd.DataFrame(columns=[
                    "问题ID", "学号", "满意程度", "问题类型", "问题内容", "答案摘要", "MCU型号", "记录时间"
                ]).to_excel(writer, sheet_name="学习记录", index=False)

                # 统计摘要表
                pd.DataFrame(columns=[
                    "学号", "提问总数", "引脚问题", "中断问题",
                    "通信问题", "编程问题", "原理问题", "其他问题",
                    "MCU_51问题", "MCU_32问题"
                ]).to_excel(writer, sheet_name="统计摘要", index=False)
            logger.info("新Excel文件创建成功")
        except Exception as e:
            logger.error(f"创建新Excel文件失败: {str(e)}")
            raise

    def _create_backup(self):
        """创建备份文件"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = os.path.join(self._backup_dir, f"学习记录_备份_{timestamp}.xlsx")
        shutil.copy2(self.excel_path, backup_path)
        return backup_path

    def _is_duplicate_record(self, df, new_record, time_window=60):
        """检查是否为重复记录"""
        current_time = pd.to_datetime(new_record["记录时间"])
        time_threshold = current_time - pd.Timedelta(seconds=time_window)

        mask = (
                (df["学号"] == new_record["学号"]) &
                (df["问题内容"] == new_record["问题内容"]) &
                (pd.to_datetime(df["记录时间"]) >= time_threshold)
        )
        return mask.any()

    def record_question(self, student_id: str, question: str, mcu_model: str, answer: Optional[str] = None) -> str:
        """记录问题（线程安全版）并返回问题ID"""
        if not student_id or not question:
            return ""

        # 生成唯一问题ID
        question_id = str(uuid.uuid4())

        with self._lock:
            try:
                if not os.path.exists(self.excel_path):
                    self._create_new_excel_file()

                backup_path = self._create_backup()
                logger.info(f"创建备份: {backup_path}")

                try:
                    df = pd.read_excel(self.excel_path, sheet_name="学习记录")
                except:
                    df = pd.DataFrame(columns=[
                        "问题ID", "学号", "满意程度", "问题类型", "问题内容", "答案摘要", "MCU型号", "记录时间"
                    ])

                new_record = {
                    "问题ID": question_id,
                    "学号": student_id,
                    "满意程度": "",
                    "问题类型": self._classify_question(question),
                    "问题内容": question,
                    "答案摘要": answer[:100] if answer else "",
                    "MCU型号": mcu_model,
                    "记录时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                }

                if not self._is_duplicate_record(df, new_record):
                    df = pd.concat([df, pd.DataFrame([new_record])], ignore_index=True)
                    self._save_and_update(df)
                    logger.info(f"问题记录成功: ID={question_id}, 学号={student_id}, 问题={question[:20]}...")
            except Exception as e:
                logger.error(f"问题记录失败: {str(e)}", exc_info=True)

        return question_id

    def _classify_question(self, question: str) -> str:
        """问题分类逻辑"""
        question = question.lower()
        if any(w in question for w in ["引脚", "管脚", "io"]):
            return "引脚问题"
        elif any(w in question for w in ["中断", "定时器"]):
            return "中断问题"
        elif any(w in question for w in ["串口", "uart", "通信"]):
            return "通信问题"
        elif any(w in question for w in ["程序", "代码"]):
            return "编程问题"
        elif any(w in question for w in ["原理", "工作"]):
            return "原理问题"
        return "其他问题"

    def _generate_summary(self, df: pd.DataFrame) -> pd.DataFrame:
        """生成统计摘要（合并同一学生的所有记录）"""
        if df.empty:
            return pd.DataFrame(columns=[
                "学号", "提问总数", "引脚问题", "中断问题",
                "通信问题", "编程问题", "原理问题", "其他问题",
                "MCU_51问题", "MCU_32问题"
            ])

        # 确保合并前先按学号分组
        summary = df.groupby('学号').agg(
            提问总数=('学号', 'size'),  # 使用size计算总提问数
            引脚问题=('问题类型', lambda x: (x == '引脚问题').sum()),
            中断问题=('问题类型', lambda x: (x == '中断问题').sum()),
            通信问题=('问题类型', lambda x: (x == '通信问题').sum()),
            编程问题=('问题类型', lambda x: (x == '编程问题').sum()),
            原理问题=('问题类型', lambda x: (x == '原理问题').sum()),
            其他问题=('问题类型', lambda x: (x == '其他问题').sum()),
            MCU_51问题=('MCU型号', lambda x: (x == '51单片机').sum()),
            MCU_32问题=('MCU型号', lambda x: (x == '32单片机').sum())
        ).reset_index()

        return summary

    def _save_and_update(self, df):
        """修复后的保存方法 - 优化性能"""
        try:
            # 仅更新变化的部分，而不是整个文件
            if not os.path.exists(self.excel_path):
                self._create_new_excel_file()

            # 使用openpyxl直接操作Excel，避免全量重写
            from openpyxl import load_workbook

            # 加载现有工作簿
            book = load_workbook(self.excel_path)

            # 更新学习记录表
            if "学习记录" in book.sheetnames:
                ws = book["学习记录"]
                # 清除旧数据（保留标题行）
                for row in range(2, ws.max_row + 1):
                    ws.delete_rows(2)

                # 添加新数据
                for _, row in df.iterrows():
                    ws.append(row.tolist())

            # 更新统计摘要表
            if "统计摘要" in book.sheetnames:
                ws = book["统计摘要"]
                # 清除旧数据（保留标题行）
                for row in range(2, ws.max_row + 1):
                    ws.delete_rows(2)

                # 生成并添加新摘要
                summary = self._generate_summary(df)
                for _, row in summary.iterrows():
                    ws.append(row.tolist())

            # 保存修改
            book.save(self.excel_path)
            logger.info("数据增量更新成功")

        except Exception as e:
            logger.error(f"数据保存失败: {str(e)}", exc_info=True)
            # 尝试恢复备份
            backups = [f for f in os.listdir(self._backup_dir) if f.endswith('.xlsx')]
            if backups:
                backups.sort(key=lambda x: os.path.getctime(os.path.join(self._backup_dir, x)), reverse=True)
                latest_backup = os.path.join(self._backup_dir, backups[0])
                try:
                    shutil.copy2(latest_backup, self.excel_path)
                    logger.info(f"已恢复备份: {latest_backup}")
                except Exception as backup_error:
                    logger.error(f"备份恢复失败: {str(backup_error)}")

    def record_rating(self, question_id: str, question: str, answer: str, rating: int) -> bool:
        """通过问题ID记录评分"""
        if not question_id:
            return False

        with self._lock:
            try:
                if not os.path.exists(self.excel_path):
                    self._create_new_excel_file()

                df = pd.read_excel(self.excel_path, sheet_name="学习记录")

                # 通过问题ID查找记录
                mask = (df["问题ID"] == question_id)

                if mask.any():
                    # 更新评分
                    df.loc[mask, "满意程度"] = rating

                    # 如果是5分，保存为独立JSON文件
                    if rating == 5:
                        # 检查是否已存在
                        if not self.top_question_manager.get_question_by_id(question_id):
                            success = self.top_question_manager.save_question(question_id, question, answer)
                            if not success:
                                logger.error(f"保存高分问题失败: ID={question_id}")
                            else:
                                logger.info(f"高分问题保存成功: ID={question_id}")

                    self._save_and_update(df)
                    logger.info(f"评分记录成功: ID={question_id}, 评分={rating}")
                    return True
                else:
                    logger.warning(f"未找到匹配的记录进行评分: ID={question_id}")
                    return False
            except Exception as e:
                logger.error(f"评分记录失败: {str(e)}", exc_info=True)
                return False

    def get_top_questions(self) -> List[Tuple[str, str, str]]:
        """获取高分问题列表（问题ID, 问题内容, 答案）"""
        return [(q["id"], q["question"], q["answer"]) for q in self.top_question_manager.get_all_questions()]

    def show_dashboard(self):
        """统计面板显示"""
        try:
            with self._lock:
                if not os.path.exists(self.excel_path):
                    self._create_new_excel_file()

                try:
                    summary_df = pd.read_excel(self.excel_path, sheet_name="统计摘要")
                    detail_df = pd.read_excel(self.excel_path, sheet_name="学习记录")
                except:
                    summary_df = pd.DataFrame()
                    detail_df = pd.DataFrame()

                st.title("📊 学情统计看板")

                tab1, tab2 = st.tabs(["统计摘要", "详细记录"])

                with tab1:
                    if not summary_df.empty:
                        sort_col = st.selectbox("排序依据", summary_df.columns[1:], index=0)
                        st.dataframe(
                            summary_df.sort_values(sort_col, ascending=False),
                            column_config={
                                "学号": "学生ID",
                                "提问总数": st.column_config.NumberColumn("提问总数"),
                                "MCU_51问题": "51单片机问题",
                                "MCU_32问题": "32单片机问题"
                            },
                            width='stretch',
                            height=600
                        )
                    else:
                        st.warning("没有统计摘要数据")

                with tab2:
                    if not detail_df.empty:
                        col1, col2 = st.columns(2)
                        with col1:
                            student_filter = st.selectbox(
                                "按学号筛选",
                                ["全部"] + list(detail_df["学号"].unique())
                            )
                        with col2:
                            mcu_filter = st.selectbox(
                                "按MCU型号筛选",
                                ["全部"] + list(detail_df["MCU型号"].unique())
                            )

                        filtered_df = detail_df
                        if student_filter != "全部":
                            filtered_df = filtered_df[filtered_df["学号"] == student_filter]
                        if mcu_filter != "全部":
                            filtered_df = filtered_df[filtered_df["MCU型号"] == mcu_filter]

                        st.dataframe(
                            filtered_df,
                            hide_index=True,
                            use_container_width=True,
                            height=600
                        )
                    else:
                        st.warning("没有详细记录数据")

        except Exception as e:
            st.error(f"加载数据失败: {str(e)}")


# 全局统计实例
tracker = ExcelTracker()


# ==================== 状态管理系统 ====================
class AppState:
    @staticmethod
    def init():
        """初始化所有会话状态"""
        if 'app_state' not in st.session_state:
            st.session_state.app_state = {
                'is_admin': False,
                'show_stats': False,
                'switch_time': time.time(),
                'memory': ConversationBufferMemory(
                    return_messages=True,
                    memory_key="chat_history",
                    output_key="answer"
                ),
                'messages': [],
                'mcu_model': "51单片机",
                'waiting_for_rating': False,
                'last_question': '',
                'last_answer': '',
                'last_question_id': '',
                'load_example': False,  # 标记是否加载了范例
                'example_question': '',  # 存储范例问题
            }
            st.session_state.app_state['memory'].chat_memory.add_message(
                AIMessage(content="你好！我是单片机助手，请问你有什么问题？")
            )

    @staticmethod
    def update(**kwargs):
        """更新状态并记录切换时间"""
        st.session_state.app_state.update({
            **kwargs,
            'switch_time': time.time()
        })


AppState.init()

# ==================== 性能优化CSS ====================
st.markdown("""
    <style>
    /* 界面切换动画 */
    .stApp {
        transition: opacity 0.3s ease;
    }
    /* 按钮响应优化 */
    .stButton>button {
        transition: transform 0.1s ease;
    }
    .stButton>button:hover {
        transform: scale(1.02);
    }
    /* 输入框优化 */
    .stTextInput>div>div>input {
        padding: 8px 12px !important;
    }
    /* 侧边栏优化 */
    .sidebar .stButton>button {
        width: 100%;
    }
    /* 范例列表样式 */
    .example-item {
        padding: 10px;
        border-radius: 5px;
        margin-bottom: 5px;
        cursor: pointer;
    }
    .example-item:hover {
        background-color: #f0f2f6;
    }
    </style>
""", unsafe_allow_html=True)


# ==================== 核心功能组件 ====================
def load_image(image_name: str) -> Image.Image:
    """优化后的图片加载函数"""
    paths = [image_name, f"images/{image_name}", f"static/{image_name}"]
    for path in paths:
        try:
            if os.path.exists(path):
                return Image.open(path)
        except Exception as e:
            logger.error(f"图片加载失败 {path}: {e}")
    return None


@st.cache_resource(ttl=3600)
def init_llm(api_key: str) -> QwenChat:
    """带缓存的LLM初始化"""
    return QwenChat(api_key=api_key, model_name="qwen-plus")


@st.cache_resource(ttl=3600)
def init_embeddings(api_key: str) -> QwenEmbeddings:
    """带缓存的Embeddings初始化"""
    return QwenEmbeddings(api_key=api_key)


# ==================== 界面组件 ====================
def show_config_panel():
    """配置面板组件"""
    with st.expander("🔧 配置选项", expanded=False):
        # 单片机型号选择
        st.markdown("选择单片机型号:")
        mcu_col1, mcu_col2 = st.columns(2)

        with mcu_col1:
            container = st.container()
            img_51 = load_image("51.jpg")
            if img_51:
                container.image(img_51, width=150, caption="51单片机")
            else:
                container.warning("51单片机图片未找到")
            if container.button("51单片机", key="btn_51"):
                AppState.update(mcu_model="51单片机")
                st.rerun()

        with mcu_col2:
            container = st.container()
            img_32 = load_image("32.jpg")
            if img_32:
                container.image(img_32, width=150, caption="32单片机")
            else:
                container.warning("32单片机图片未找到")
            if container.button("32单片机", key="btn_32"):
                AppState.update(mcu_model="32单片机")
                st.rerun()

        st.markdown(f"**当前选择**: {st.session_state.app_state['mcu_model']}")

        if st.button("🔄 重置对话"):
            AppState.update(
                memory=ConversationBufferMemory(
                    return_messages=True,
                    memory_key="chat_history",
                    output_key="answer"
                ),
                messages=[]
            )
            st.session_state.app_state['memory'].chat_memory.add_message(
                AIMessage(content="对话已重置，请重新提问。")
            )
            st.rerun()


def show_admin_panel():
    """管理员控制面板"""
    # 在侧边栏最下方添加管理员登录按钮
    st.markdown("---")  # 添加分隔线

    # 非管理员状态显示登录按钮
    if not st.session_state.app_state['is_admin']:
        # 使用空列将按钮推到最下方
        for _ in range(60):  # 添加多个空行
            st.write("")

        # 添加管理员登录按钮
        if st.button("🔑 管理员登录", use_container_width=True):
            # 设置状态显示密码输入框
            st.session_state.show_admin_password = True

    # 如果用户点击了管理员登录按钮，显示密码输入框
    if st.session_state.get('show_admin_password', False):
        password = st.text_input("请输入管理员密码", type="password", key="admin_pw")

        if password:
            if password == "qwert":
                AppState.update(is_admin=True, show_stats=True)
                # 重置状态
                st.session_state.show_admin_password = False
                st.rerun()
            else:
                st.error("密码错误")

    # 管理员状态处理
    if st.session_state.app_state['is_admin']:
        st.success("管理员模式")
        col1, col2 = st.columns(2)
        with col1:
            if st.button("📊 统计面板", use_container_width=True):
                AppState.update(show_stats=True)
        with col2:
            if st.button("❓ 问答界面", use_container_width=True):
                AppState.update(show_stats=False)
        if st.button("🚪 退出管理", type="primary", use_container_width=True):
            AppState.update(is_admin=False, show_stats=False)
            st.rerun()


# ==================== 主界面路由 ====================
def show_stats_interface():
    """统计面板界面"""
    with st.spinner("正在加载数据..."):
        tracker.show_dashboard()
    if st.button("← 返回问答界面"):
        AppState.update(show_stats=False)


def show_rating_buttons(question_id: str, question: str, answer: str, student_id: str):
    """显示评分按钮组件"""
    if not student_id:
        st.warning("请先输入学号以进行评分")
        return

    st.markdown("---")
    st.markdown("#### 💡 请为这个回答评分:")

    # 使用columns创建并排的评分按钮
    col1, col2, col3, col4, col5 = st.columns(5)

    # 为每个评分按钮设置不同的标签和值
    rating_options = [
        {"label": "⭐", "value": 1, "help": "评分: 1分 (完全不满意)"},
        {"label": "⭐⭐", "value": 2, "help": "评分: 2分 (不太满意)"},
        {"label": "⭐⭐⭐", "value": 3, "help": "评分: 3分 (一般满意)"},
        {"label": "⭐⭐⭐⭐", "value": 4, "help": "评分: 4分 (比较满意)"},
        {"label": "⭐⭐⭐⭐⭐", "value": 5, "help": "评分: 5分 (非常满意)"}
    ]

    # 创建5个评分按钮
    for i, col in enumerate([col1, col2, col3, col4, col5]):
        with col:
            if st.button(
                    rating_options[i]["label"],
                    help=rating_options[i]["help"],
                    use_container_width=True,
                    key=f"rating_{i + 1}_{question_id}"  # 添加唯一标识符
            ):
                # 显示加载指示器
                with st.spinner("正在记录评分..."):
                    success = tracker.record_rating(question_id, question, answer, rating_options[i]["value"])

                if success:
                    st.success(f"感谢您的评分！({rating_options[i]['value']}分)")
                    st.session_state.app_state['waiting_for_rating'] = False
                    # 使用st.rerun()替代st.experimental_rerun()
                    st.rerun()
                else:
                    st.error("评分失败，请稍后再试")

    # 添加跳过评分选项
    if st.button("跳过评分", use_container_width=True, key=f"skip_{question_id}"):
        st.session_state.app_state['waiting_for_rating'] = False
        st.rerun()


# ==================== 主界面路由 ====================
def show_qa_interface():
    """主问答界面"""
    student_id = st.session_state.get('student_id', '')
    if student_id:
        st.info(f"当前学号: {student_id}")

    # 检查是否需要加载范例问题
    if st.session_state.app_state.get('load_example', False):
        example_question = st.session_state.app_state['example_question']
        example_answer = st.session_state.app_state.get('example_answer', '')

        # 将范例问题添加到消息列表
        st.session_state.app_state['messages'].append(
            {"role": "user", "content": example_question}
        )

        # 将范例答案添加到消息列表
        st.session_state.app_state['messages'].append(
            {"role": "assistant", "content": example_answer}
        )

        # 添加到记忆
        memory = st.session_state.app_state['memory']
        memory.chat_memory.add_message(HumanMessage(content=example_question))
        memory.chat_memory.add_message(AIMessage(content=example_answer))

        # 记录问题（如果需要）
        if student_id:
            question_id = tracker.record_question(
                student_id=student_id,
                question=example_question,
                mcu_model=st.session_state.app_state['mcu_model'],
                answer=example_answer[:100]
            )

            if question_id:
                st.session_state.app_state.update({
                    'last_question': example_question,
                    'last_answer': example_answer,
                    'last_question_id': question_id,
                    'waiting_for_rating': True
                })

        # 重置状态
        st.session_state.app_state.update({
            'load_example': False,
            'example_question': '',
            'example_answer': ''
        })
        st.rerun()

    # 处理用户输入的问题
    question = st.chat_input("输入您的问题...")

    # 检查是否需要加载范例问题 - 修复加载机制
    if st.session_state.app_state.get('load_example', False):
        # 直接设置问题变量，避免JS注入
        question = st.session_state.app_state['example_question']
        # 重置状态
        st.session_state.app_state.update({
            'load_example': False,
            'example_question': ''
        })

    if question:
        # 先将用户的问题添加到消息列表
        st.session_state.app_state['messages'].append(
            {"role": "user", "content": question}
        )

        st.session_state.app_state['waiting_for_rating'] = True
        st.session_state.pop('rating_submitted', None)
        api_key = st.session_state.get('api_key', '')
        if not api_key:
            st.error("请先输入API密钥")
            return

        with st.spinner("正在思考..."):
            try:
                llm = init_llm(api_key)
                embeddings = init_embeddings(api_key)

                result = qa_agent(
                    question=question,
                    memory=st.session_state.app_state['memory'],
                    mcu_model=st.session_state.app_state['mcu_model'],
                    llm=llm,
                    embeddings=embeddings
                )

                answer = result.get("answer", "抱歉，我没有找到答案。")
                st.session_state.app_state['messages'].append(
                    {"role": "assistant", "content": answer}
                )

                if student_id:
                    # 记录问题并获取问题ID
                    question_id = tracker.record_question(
                        student_id=student_id,
                        question=question,
                        mcu_model=st.session_state.app_state['mcu_model'],
                        answer=answer[:100]
                    )

                    if question_id:
                        st.session_state.app_state.update({
                            'last_question': question,
                            'last_answer': answer,
                            'last_question_id': question_id,
                            'waiting_for_rating': True
                        })
                    else:
                        st.error("问题记录失败，无法评分")
                        AppState.update(waiting_for_rating=False)

            except Exception as e:
                error_msg = f"系统错误: {str(e)}"
                st.error(error_msg)
                logger.error(traceback.format_exc())

    # 显示最近的10条消息
    messages = st.session_state.app_state.get('messages', [])
    for msg in messages[-10:]:
        with st.chat_message(msg["role"]):
            st.markdown(msg["content"])

    # 显示评分按钮（如果需要）
    if (st.session_state.app_state.get('waiting_for_rating', False) and
            student_id and
            st.session_state.app_state.get('last_question') and
            st.session_state.app_state.get('last_question_id')):
        show_rating_buttons(
            st.session_state.app_state['last_question_id'],
            st.session_state.app_state['last_question'],
            st.session_state.app_state['last_answer'],
            student_id
        )


# ==================== 提问范例组件 ====================
def show_examples_panel():
    """修复后的提问范例面板"""
    with st.expander("💡 提问范例", expanded=False):
        st.info("以下是一些高质量的提问范例（用户评分5分的问题）：")

        top_questions = tracker.get_top_questions()
        if not top_questions:
            st.warning("暂无高质量问题范例")
            return

        # 显示问题列表 - 使用Streamlit原生按钮
        for i, (qid, question, answer) in enumerate(top_questions):
            # 显示简略问题文本
            truncated_question = question[:50] + ('...' if len(question) > 50 else '')

            # 使用Streamlit按钮替代JS点击事件
            if st.button(
                    f"范例 {i + 1}: {truncated_question}",
                    key=f"example_btn_{i}",
                    use_container_width=True
            ):
                # 直接更新状态，不需要JS
                st.session_state.app_state.update({
                    'load_example': True,
                    'example_question': question,
                    'example_answer': answer  # 存储范例答案
                })
                st.rerun()


# ==================== 应用主入口 ====================
def main():
    # 在侧边栏组织所有用户配置选项
    with st.sidebar:
        st.header("配置选项")

        # API密钥和学号输入
        st.session_state['api_key'] = st.text_input(
            "通义千问API密钥",
            type="password",
            placeholder="输入API密钥",
            key="api_key_input"
        )
        st.session_state['student_id'] = st.text_input(
            "学号",
            placeholder="请输入学号",
            key="student_id_input"
        )

        # 侧边栏重置对话按钮
        if st.button("🔄 重置对话", use_container_width=True, key="reset_sidebar"):
            AppState.update(
                memory=ConversationBufferMemory(
                    return_messages=True,
                    memory_key="chat_history",
                    output_key="answer"
                ),
                messages=[]
            )
            st.session_state.app_state['memory'].chat_memory.add_message(
                AIMessage(content="对话已重置，请重新提问。")
            )
            st.rerun()

        show_examples_panel()

        # 在侧边栏显示管理员面板
        show_admin_panel()

        st.markdown("---")  # 分隔线

    # 根据状态显示不同界面
    if st.session_state.app_state['is_admin'] and st.session_state.app_state['show_stats']:
        show_stats_interface()
    else:
        # 在主内容区显示配置面板（放在标题下方）
        st.title("🖥️ 单片机智能问答工具")
        show_config_panel()  # 将配置面板放在标题下方
        show_qa_interface()


if __name__ == "__main__":
    main()
